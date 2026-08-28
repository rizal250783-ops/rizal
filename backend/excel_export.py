"""Excel export for notes (openpyxl) - styled & colored."""
import io
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

TEAL = "00A0A0"
TEAL_DARK = "0F766E"
GOLD = "F0B43C"
GOLD_LIGHT = "FDF7EB"
ZEBRA = "F1FAFA"
WHITE = "FFFFFF"

STATUS_FILL = {
    "Final Approved": "D1FAE5",
    "Draft": "F1F5F9",
}


def _status_fill(status: str) -> str:
    if not status:
        return WHITE
    if status in STATUS_FILL:
        return STATUS_FILL[status]
    if status.startswith("Menunggu"):
        return "FDF7EB"
    if status.startswith("Revisi"):
        return "FEF3C7"
    if status.startswith("Reject") or status.startswith("Memerlukan"):
        return "FEE2E2"
    return WHITE


def export_notes_excel(notes: list, meta: dict = None) -> bytes:
    meta = meta or {}
    wb = Workbook()
    ws = wb.active
    ws.title = "Nota Restruktur"

    headers = ["No", "Nomor Nota", "Nama Nasabah", "CIF", "Cabang", "Area", "Region", "Segmen", "Status",
               "Total OS Pokok", "Total OS Margin", "Total Penalty",
               "Total Kewajiban", "Pemutus", "Tanggal Approved"]
    ncol = len(headers)
    last_col = get_column_letter(ncol)

    thin = Side(style="thin", color="CBD5E1")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Row 1: Title band
    ws.merge_cells(f"A1:{last_col}1")
    t = ws["A1"]
    t.value = "DAFTAR NOTA ANALISA RESTRUKTUR PEMBIAYAAN — PT. Bank Syariah Indonesia, Tbk"
    t.font = Font(bold=True, size=14, color=WHITE)
    t.fill = PatternFill("solid", fgColor=TEAL)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 26

    # Row 2: Subtitle / meta band
    ws.merge_cells(f"A2:{last_col}2")
    s = ws["A2"]
    filt = meta.get("filter_text") or "Semua data (sesuai kewenangan)"
    s.value = (f"Diekspor: {datetime.now().strftime('%d/%m/%Y %H:%M')}   |   "
               f"Oleh: {meta.get('by', '-')}   |   Filter: {filt}   |   Jumlah: {len(notes)} nota")
    s.font = Font(size=9, italic=True, color="0F172A")
    s.fill = PatternFill("solid", fgColor=GOLD_LIGHT)
    s.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[2].height = 18

    # Row 3: Header
    header_row = 3
    ws.append([]) if False else None
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=header_row, column=i, value=h)
        c.font = Font(bold=True, color=WHITE, size=10)
        c.fill = PatternFill("solid", fgColor=TEAL_DARK)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = border
    ws.row_dimensions[header_row].height = 30

    money_cols = (10, 11, 12, 13)
    r = header_row
    for idx, n in enumerate(notes, 1):
        r += 1
        cust = n.get("customer", {})
        facs = n.get("facilities", [])
        cif = facs[0].get("cif", "") if facs else ""
        cabang = ", ".join(sorted({f.get("nama_cabang", "") for f in facs if f.get("nama_cabang")}))
        segmen = ", ".join(sorted({f.get("segmen", "") for f in facs if f.get("segmen")}))
        row_vals = [
            idx, n.get("nomor_nota", ""), cust.get("nama", ""), cif, cabang, n.get("area", ""), n.get("region", ""),
            segmen, n.get("status", ""),
            n.get("total_os_pokok", 0), n.get("total_os_margin", 0), n.get("total_penalty", 0),
            n.get("total_kewajiban", 0), n.get("final_approver_nama", "") or "-",
            n.get("approved_date", "") or "-",
        ]
        zebra = (idx % 2 == 0)
        for ci, val in enumerate(row_vals, 1):
            c = ws.cell(row=r, column=ci, value=val)
            c.border = border
            c.font = Font(size=9)
            if zebra:
                c.fill = PatternFill("solid", fgColor=ZEBRA)
            if ci in money_cols:
                c.number_format = '#,##0'
                c.alignment = Alignment(horizontal="right")
            elif ci == 1:
                c.alignment = Alignment(horizontal="center")
        # status color overrides zebra
        sc = ws.cell(row=r, column=9)
        sc.fill = PatternFill("solid", fgColor=_status_fill(n.get("status", "")))
        sc.alignment = Alignment(horizontal="center")
        sc.font = Font(size=9, bold=True)

    # Totals row
    if notes:
        r += 1
        tot = ws.cell(row=r, column=1, value="TOTAL")
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=9)
        tot.font = Font(bold=True, color="0F172A")
        tot.fill = PatternFill("solid", fgColor=GOLD)
        tot.alignment = Alignment(horizontal="right", vertical="center")
        sums = {
            10: sum(float(n.get("total_os_pokok", 0) or 0) for n in notes),
            11: sum(float(n.get("total_os_margin", 0) or 0) for n in notes),
            12: sum(float(n.get("total_penalty", 0) or 0) for n in notes),
            13: sum(float(n.get("total_kewajiban", 0) or 0) for n in notes),
        }
        for col in range(1, ncol + 1):
            cc = ws.cell(row=r, column=col)
            cc.border = border
            if col in sums:
                cc.value = sums[col]
                cc.number_format = '#,##0'
                cc.font = Font(bold=True, color="0F172A")
                cc.fill = PatternFill("solid", fgColor=GOLD)
                cc.alignment = Alignment(horizontal="right")
            elif col > 1:
                cc.fill = PatternFill("solid", fgColor=GOLD)

    widths = [5, 24, 24, 14, 26, 20, 20, 14, 22, 18, 18, 16, 20, 24, 16]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = f"A{header_row + 1}"
    ws.auto_filter.ref = f"A{header_row}:{last_col}{header_row}"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
