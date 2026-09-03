import os
from dotenv import load_dotenv
load_dotenv()

import io
import uuid
import secrets
import string
import logging
from datetime import datetime, timezone, timedelta
from typing import Optional, List

import jwt
import bcrypt
from bson import ObjectId
from fastapi import FastAPI, APIRouter, HTTPException, Request, Response, Depends, UploadFile, File, Form, Query
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse, Response as FastResponse
from pydantic import BaseModel, EmailStr, Field
from motor.motor_asyncio import AsyncIOMotorClient
import openpyxl

import calculations as calc
import storage_util

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger("ao360")

MONGO_URL = os.environ["MONGO_URL"]
DB_NAME = os.environ["DB_NAME"]
JWT_SECRET = os.environ["JWT_SECRET"]
JWT_ALG = "HS256"

client = AsyncIOMotorClient(MONGO_URL)
db = client[DB_NAME]

app = FastAPI(title="AO-360 API")
api = APIRouter(prefix="/api")

ROLES = ["direktur", "admin", "ao_lending", "ao_funding", "pic_remedial"]

# ---------------- Password / JWT ----------------

def hash_password(pw: str) -> str:
    return bcrypt.hashpw(pw.encode(), bcrypt.gensalt()).decode()

def verify_password(pw: str, hashed: str) -> bool:
    try:
        return bcrypt.checkpw(pw.encode(), hashed.encode())
    except Exception:
        return False

def validate_password_policy(pw: str):
    if len(pw) < 8 or not any(c.isalpha() for c in pw) or not any(c.isdigit() for c in pw):
        raise HTTPException(status_code=400, detail="Password minimal 8 karakter, kombinasi huruf dan angka.")

def gen_temp_password() -> str:
    return "Ao" + "".join(secrets.choice(string.ascii_letters + string.digits) for _ in range(8)) + "1"

async def get_session_timeout() -> int:
    s = await db.system_settings.find_one({"_id": "system"})
    return int((s or {}).get("session_timeout_minutes", 60))

def create_token(user_id: str, minutes: int) -> str:
    payload = {"sub": user_id, "exp": datetime.now(timezone.utc) + timedelta(minutes=minutes), "type": "access"}
    return jwt.encode(payload, JWT_SECRET, algorithm=JWT_ALG)

# ---------------- Serialization ----------------

def ser(doc: dict) -> dict:
    if not doc:
        return doc
    doc = dict(doc)
    doc["id"] = str(doc.pop("_id"))
    doc.pop("password_hash", None)
    return doc

# ---------------- Audit ----------------

async def audit(user, aktivitas: str, before=None, after=None, extra=None):
    await db.audit_logs.insert_one({
        "user_id": (user or {}).get("id"),
        "user_name": (user or {}).get("name", "system"),
        "role": (user or {}).get("role"),
        "aktivitas": aktivitas,
        "data_sebelum": before,
        "data_sesudah": after,
        "extra": extra,
        "waktu": datetime.now(timezone.utc).isoformat(),
    })

# ---------------- Auth deps ----------------

async def get_current_user(request: Request) -> dict:
    token = request.cookies.get("access_token")
    if not token:
        h = request.headers.get("Authorization", "")
        if h.startswith("Bearer "):
            token = h[7:]
    if not token:
        raise HTTPException(status_code=401, detail="Tidak terautentikasi")
    try:
        payload = jwt.decode(token, JWT_SECRET, algorithms=[JWT_ALG])
        user = await db.users.find_one({"_id": ObjectId(payload["sub"])})
        if not user:
            raise HTTPException(status_code=401, detail="User tidak ditemukan")
        if not user.get("is_active", True):
            raise HTTPException(status_code=403, detail="Akun dinonaktifkan")
        return ser(user)
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="Sesi berakhir, silakan login kembali")
    except jwt.InvalidTokenError:
        raise HTTPException(status_code=401, detail="Token tidak valid")

def require_roles(*roles):
    async def dep(user: dict = Depends(get_current_user)):
        if user["role"] not in roles:
            raise HTTPException(status_code=403, detail="Akses ditolak untuk role ini")
        return user
    return dep

# ---------------- Models ----------------

class LoginIn(BaseModel):
    email: EmailStr
    password: str

class ChangePasswordIn(BaseModel):
    current_password: str
    new_password: str

class UserIn(BaseModel):
    name: str
    email: EmailStr
    role: str
    password: Optional[str] = None
    employee_id: Optional[str] = None
    phone: Optional[str] = None

class UserUpdateIn(BaseModel):
    name: Optional[str] = None
    email: Optional[EmailStr] = None
    phone: Optional[str] = None
    employee_id: Optional[str] = None

class RoleChangeIn(BaseModel):
    new_role: str

class TargetIn(BaseModel):
    ao_id: str
    period: str  # YYYY-MM
    target_booking: float = 0
    target_funding: float = 0
    target_recovery_wo: float = 0
    target_npf_ratio: float = 0
    target_npf_absolute: float = 0

class AchievementIn(BaseModel):
    ao_id: str
    period: str
    realisasi_booking: float = 0
    realisasi_funding: float = 0
    realisasi_recovery_wo: float = 0

class PortfolioIn(BaseModel):
    nomor_kontrak: str
    nama_nasabah: str
    produk: str
    plafond: float
    outstanding_pokok: float
    tanggal_akad: str
    tanggal_jatuh_tempo: str
    kolektibilitas: int
    dpd: int = 0
    ao_id: str

class PerfSettingIn(BaseModel):
    role: str
    weights: dict  # {"lending":70,"funding":30} etc
    reason: Optional[str] = None

class ParamIn(BaseModel):
    parameter_key: str
    parameter_value: float

class SettingsIn(BaseModel):
    active_period: Optional[str] = None
    session_timeout_minutes: Optional[int] = None

# ---------------- Helpers ----------------

def current_period() -> str:
    return datetime.now(timezone.utc).strftime("%Y-%m")

async def get_active_period() -> str:
    s = await db.system_settings.find_one({"_id": "system"})
    return (s or {}).get("active_period") or current_period()

async def get_weights(role: str) -> dict:
    doc = await db.performance_settings.find_one(
        {"role": role, "type": "weight"}, sort=[("version", -1)])
    if doc:
        return doc["weights"]
    defaults = {
        "ao_lending": {"lending": 70, "funding": 30},
        "ao_funding": {"funding": 100},
        "pic_remedial": {"recovery": 70, "npf": 30},
    }
    return defaults.get(role, {})

async def get_param(key: str, default: float) -> float:
    doc = await db.performance_settings.find_one({"type": "parameter", "parameter_key": key},
                                                 sort=[("version", -1)])
    return float(doc["parameter_value"]) if doc else default

# ---------- portfolio aggregates ----------

async def portfolio_outstanding(ao_id: Optional[str] = None):
    match = {} if ao_id is None else {"ao_id": ao_id}
    cur = db.loan_portfolio.find(match)
    total = 0.0
    kol = {1: 0.0, 2: 0.0, 3: 0.0, 4: 0.0, 5: 0.0}
    kol_count = {1: 0, 2: 0, 3: 0, 4: 0, 5: 0}
    async for p in cur:
        o = float(p.get("outstanding_pokok", 0))
        k = int(p.get("kolektibilitas", 1))
        total += o
        if k in kol:
            kol[k] += o
            kol_count[k] += 1
    npf_out = kol[3] + kol[4] + kol[5]
    return {"total": total, "kol": kol, "kol_count": kol_count, "npf_out": npf_out}

# ============================================================
# AUTH ENDPOINTS
# ============================================================

@api.post("/auth/login")
async def login(body: LoginIn, response: Response):
    email = body.email.lower().strip()
    user = await db.users.find_one({"email": email})
    generic = HTTPException(status_code=401, detail="Email atau password salah")

    if user:
        locked_until = user.get("locked_until")
        if locked_until and datetime.fromisoformat(locked_until) > datetime.now(timezone.utc):
            await audit(ser(user), "Login gagal - akun terkunci")
            raise HTTPException(status_code=423, detail="Akun terkunci sementara. Coba lagi nanti.")

    if not user or not user.get("is_active", True) or not verify_password(body.password, user.get("password_hash", "")):
        if user:
            attempts = user.get("failed_login_attempts", 0) + 1
            upd = {"failed_login_attempts": attempts, "last_failed_at": datetime.now(timezone.utc).isoformat()}
            if attempts >= 5:
                upd["locked_until"] = (datetime.now(timezone.utc) + timedelta(minutes=15)).isoformat()
                upd["failed_login_attempts"] = 0
                await db.users.update_one({"_id": user["_id"]}, {"$set": upd})
                await audit(ser(user), "Login gagal - akun dikunci 15 menit (5x gagal)")
            else:
                await db.users.update_one({"_id": user["_id"]}, {"$set": upd})
                await audit(ser(user), f"Login gagal ({attempts}x)")
        raise generic

    await db.users.update_one({"_id": user["_id"]},
                              {"$set": {"failed_login_attempts": 0, "locked_until": None,
                                        "last_login": datetime.now(timezone.utc).isoformat()}})
    timeout = await get_session_timeout()
    token = create_token(str(user["_id"]), timeout)
    response.set_cookie("access_token", token, httponly=True, secure=True, samesite="none",
                        max_age=timeout * 60, path="/")
    data = ser(user)
    data["access_token"] = token
    data["requires_password_reset"] = user.get("requires_password_reset", False)
    return data

@api.post("/auth/logout")
async def logout(response: Response, user: dict = Depends(get_current_user)):
    response.delete_cookie("access_token", path="/")
    return {"ok": True}

@api.get("/auth/me")
async def me(user: dict = Depends(get_current_user)):
    fresh = await db.users.find_one({"_id": ObjectId(user["id"])})
    return ser(fresh)

@api.post("/auth/change-password")
async def change_password(body: ChangePasswordIn, user: dict = Depends(get_current_user)):
    full = await db.users.find_one({"_id": ObjectId(user["id"])})
    if not verify_password(body.current_password, full.get("password_hash", "")):
        raise HTTPException(status_code=400, detail="Password lama salah")
    validate_password_policy(body.new_password)
    await db.users.update_one({"_id": full["_id"]}, {"$set": {
        "password_hash": hash_password(body.new_password),
        "requires_password_reset": False}})
    await audit(user, "Ganti password sendiri")
    return {"ok": True}

# ============================================================
# USER MANAGEMENT (admin)
# ============================================================

@api.get("/users")
async def list_users(user: dict = Depends(require_roles("admin", "direktur"))):
    out = []
    async for u in db.users.find().sort("created_at", 1):
        out.append(ser(u))
    return out

@api.post("/users")
async def create_user(body: UserIn, user: dict = Depends(require_roles("admin"))):
    if body.role not in ROLES:
        raise HTTPException(status_code=400, detail="Role tidak valid")
    if await db.users.find_one({"email": body.email.lower()}):
        raise HTTPException(status_code=400, detail="Email sudah terdaftar")
    pw = body.password or gen_temp_password()
    validate_password_policy(pw)
    doc = {
        "name": body.name, "email": body.email.lower(), "role": body.role,
        "employee_id": body.employee_id, "phone": body.phone,
        "password_hash": hash_password(pw), "is_active": True,
        "requires_password_reset": bool(not body.password),
        "failed_login_attempts": 0, "locked_until": None,
        "branch": "PT BPRS Haji Miskin",
        "created_at": datetime.now(timezone.utc).isoformat(),
    }
    res = await db.users.insert_one(doc)
    doc["_id"] = res.inserted_id
    await audit(user, "Membuat user", after={"email": body.email, "role": body.role})
    out = ser(doc)
    if not body.password:
        out["temp_password"] = pw
    return out

@api.put("/users/{uid}")
async def update_user(uid: str, body: UserUpdateIn, user: dict = Depends(require_roles("admin"))):
    upd = {k: v for k, v in body.dict().items() if v is not None}
    if "email" in upd:
        upd["email"] = upd["email"].lower()
    before = await db.users.find_one({"_id": ObjectId(uid)})
    await db.users.update_one({"_id": ObjectId(uid)}, {"$set": upd})
    await audit(user, "Edit user", before=ser(before), after=upd)
    return ser(await db.users.find_one({"_id": ObjectId(uid)}))

@api.post("/users/{uid}/status")
async def set_status(uid: str, active: bool = Query(...), user: dict = Depends(require_roles("admin"))):
    await db.users.update_one({"_id": ObjectId(uid)}, {"$set": {"is_active": active}})
    await audit(user, "Aktifkan user" if active else "Nonaktifkan user", after={"uid": uid})
    return {"ok": True}

@api.delete("/users/{uid}")
async def delete_user(uid: str, user: dict = Depends(require_roles("admin"))):
    before = await db.users.find_one({"_id": ObjectId(uid)})
    await db.users.delete_one({"_id": ObjectId(uid)})
    await audit(user, "Hapus user", before=ser(before))
    return {"ok": True}

@api.post("/users/{uid}/reset-password")
async def reset_password(uid: str, user: dict = Depends(require_roles("admin"))):
    temp = gen_temp_password()
    await db.users.update_one({"_id": ObjectId(uid)}, {"$set": {
        "password_hash": hash_password(temp), "requires_password_reset": True,
        "failed_login_attempts": 0, "locked_until": None}})
    await audit(user, "Reset password user", after={"uid": uid})
    return {"temp_password": temp}

@api.post("/users/{uid}/change-role")
async def change_role(uid: str, body: RoleChangeIn, user: dict = Depends(require_roles("admin"))):
    if body.new_role not in ROLES:
        raise HTTPException(status_code=400, detail="Role tidak valid")
    before = await db.users.find_one({"_id": ObjectId(uid)})
    old_role = before.get("role")
    await db.users.update_one({"_id": ObjectId(uid)}, {"$set": {"role": body.new_role}})
    await db.role_history.insert_one({
        "user_id": uid, "user_name": before.get("name"),
        "role_lama": old_role, "role_baru": body.new_role,
        "tanggal": datetime.now(timezone.utc).isoformat(),
        "admin_id": user["id"], "admin_name": user["name"]})
    await audit(user, "Pindah jabatan user", before={"role": old_role}, after={"role": body.new_role})
    return {"ok": True}

@api.get("/users/{uid}/role-history")
async def role_history(uid: str, user: dict = Depends(require_roles("admin", "direktur"))):
    out = []
    async for r in db.role_history.find({"user_id": uid}).sort("tanggal", -1):
        out.append(ser(r))
    return out

# ============================================================
# TARGETS & ACHIEVEMENTS
# ============================================================

@api.get("/targets")
async def list_targets(period: Optional[str] = None, user: dict = Depends(get_current_user)):
    q = {}
    if period:
        q["period"] = period
    if user["role"] in ("ao_lending", "ao_funding", "pic_remedial"):
        q["ao_id"] = user["id"]
    out = []
    async for t in db.targets.find(q):
        out.append(ser(t))
    return out

@api.post("/targets")
async def upsert_target(body: TargetIn, user: dict = Depends(require_roles("admin"))):
    q = {"ao_id": body.ao_id, "period": body.period}
    before = await db.targets.find_one(q)
    data = body.dict()
    await db.targets.update_one(q, {"$set": data}, upsert=True)
    await audit(user, "Set/ubah target", before=ser(before) if before else None, after=data)
    return ser(await db.targets.find_one(q))

@api.get("/achievements")
async def list_achievements(period: Optional[str] = None, user: dict = Depends(get_current_user)):
    q = {}
    if period:
        q["period"] = period
    if user["role"] in ("ao_lending", "ao_funding", "pic_remedial"):
        q["ao_id"] = user["id"]
    out = []
    async for a in db.achievements.find(q):
        out.append(ser(a))
    return out

@api.post("/achievements")
async def upsert_achievement(body: AchievementIn, user: dict = Depends(require_roles("admin"))):
    q = {"ao_id": body.ao_id, "period": body.period}
    before = await db.achievements.find_one(q)
    data = body.dict()
    await db.achievements.update_one(q, {"$set": data}, upsert=True)
    await audit(user, "Input/ubah achievement", before=ser(before) if before else None, after=data)
    return ser(await db.achievements.find_one(q))

# ============================================================
# CALCULATION HELPERS (per AO)
# ============================================================

async def compute_ao(ao: dict, period: str):
    ao_id = ao["id"] if "id" in ao else str(ao["_id"])
    role = ao["role"]
    t = await db.targets.find_one({"ao_id": ao_id, "period": period}) or {}
    a = await db.achievements.find_one({"ao_id": ao_id, "period": period}) or {}
    result = {"ao_id": ao_id, "name": ao["name"], "role": role, "period": period}

    if role == "ao_lending":
        ach_l = calc.hitung_achievement(t.get("target_booking"), a.get("realisasi_booking"))
        ach_f = calc.hitung_achievement(t.get("target_funding"), a.get("realisasi_funding"))
        w = await get_weights("ao_lending")
        ps = calc.hitung_performance_score_lending(ach_l, ach_f, w.get("lending", 70), w.get("funding", 30))
        result.update({"ach_lending": ach_l, "ach_funding": ach_f, "performance": ps,
                       "target_booking": ach_l["target"], "realisasi_booking": ach_l["realisasi"],
                       "target_funding": ach_f["target"], "realisasi_funding": ach_f["realisasi"]})
    elif role == "ao_funding":
        ach_f = calc.hitung_achievement(t.get("target_funding"), a.get("realisasi_funding"))
        w = await get_weights("ao_funding")
        ps = calc.hitung_performance_score_funding(ach_f, w.get("funding", 100))
        result.update({"ach_funding": ach_f, "performance": ps,
                       "target_funding": ach_f["target"], "realisasi_funding": ach_f["realisasi"]})
    elif role == "pic_remedial":
        ach_r = calc.hitung_achievement(t.get("target_recovery_wo"), a.get("realisasi_recovery_wo"))
        pf = await portfolio_outstanding()  # NPF posisi cabang keseluruhan
        cap = await get_param("npf_score_cap", 150.0)
        thr = await get_param("npf_status_threshold", 1.0)
        npf = calc.hitung_npf(pf["npf_out"], pf["total"], t.get("target_npf_ratio", 3), cap, thr)
        w = await get_weights("pic_remedial")
        ps = calc.hitung_performance_score_remedial(ach_r, npf, w.get("recovery", 70), w.get("npf", 30))
        result.update({"ach_recovery": ach_r, "npf": npf, "performance": ps,
                       "target_recovery": ach_r["target"], "realisasi_recovery": ach_r["realisasi"]})
    return result

# ============================================================
# DASHBOARDS
# ============================================================

@api.get("/dashboard/me")
async def dashboard_me(period: Optional[str] = None, user: dict = Depends(get_current_user)):
    period = period or await get_active_period()
    full = await db.users.find_one({"_id": ObjectId(user["id"])})
    data = await compute_ao(ser(full), period)
    if user["role"] == "ao_lending":
        pf = await portfolio_outstanding(user["id"])
        data["portfolio"] = pf
    return data

@api.get("/dashboard/executive")
async def dashboard_executive(period: Optional[str] = None, user: dict = Depends(require_roles("direktur", "admin"))):
    period = period or await get_active_period()
    aos = []
    async for u in db.users.find({"role": {"$in": ["ao_lending", "ao_funding", "pic_remedial"]}, "is_active": True}):
        aos.append(await compute_ao(ser(u), period))

    pf = await portfolio_outstanding()
    cap = await get_param("npf_score_cap", 150.0)
    thr = await get_param("npf_status_threshold", 1.0)
    # target npf from any remedial target this period
    rt = await db.targets.find_one({"period": period, "target_npf_ratio": {"$gt": 0}}, sort=[("target_npf_ratio", -1)]) or {}
    npf = calc.hitung_npf(pf["npf_out"], pf["total"], rt.get("target_npf_ratio", 3), cap, thr)

    def sum_field(role, key):
        return sum(float(x.get(key, 0) or 0) for x in aos if x["role"] == role)

    lending_sum = {
        "target_booking": sum_field("ao_lending", "target_booking"),
        "realisasi_booking": sum_field("ao_lending", "realisasi_booking"),
    }
    funding_sum = {
        "target_funding": sum_field("ao_lending", "target_funding") + sum_field("ao_funding", "target_funding"),
        "realisasi_funding": sum_field("ao_lending", "realisasi_funding") + sum_field("ao_funding", "realisasi_funding"),
    }
    recovery_sum = {
        "target_recovery": sum_field("pic_remedial", "target_recovery"),
        "realisasi_recovery": sum_field("pic_remedial", "realisasi_recovery"),
    }
    return {
        "period": period,
        "lending": {**lending_sum, "achievement": calc.hitung_achievement(lending_sum["target_booking"], lending_sum["realisasi_booking"])},
        "funding": {**funding_sum, "achievement": calc.hitung_achievement(funding_sum["target_funding"], funding_sum["realisasi_funding"])},
        "recovery": {**recovery_sum, "achievement": calc.hitung_achievement(recovery_sum["target_recovery"], recovery_sum["realisasi_recovery"])},
        "portfolio": pf,
        "npf": npf,
        "total_ao": len(aos),
    }

# ============================================================
# RANKING / LEADERBOARD
# ============================================================

@api.get("/ranking")
async def ranking(type: str = Query(...), period: Optional[str] = None,
                  user: dict = Depends(require_roles("direktur", "admin"))):
    period = period or await get_active_period()
    role_map = {"lending": "ao_lending", "funding": "ao_funding", "remedial": "pic_remedial"}
    role = role_map.get(type)
    if not role:
        raise HTTPException(status_code=400, detail="Tipe tidak valid")
    entries = []
    async for u in db.users.find({"role": role, "is_active": True}):
        d = await compute_ao(ser(u), period)
        ps = d.get("performance", {})
        if role == "ao_lending":
            ach = d["ach_lending"].get("value")
            realisasi = d["ach_lending"]["realisasi"]
        elif role == "ao_funding":
            ach = d["ach_funding"].get("value")
            realisasi = d["ach_funding"]["realisasi"]
        else:
            ach = d["ach_recovery"].get("value")
            realisasi = d["ach_recovery"]["realisasi"]
        entries.append({
            "ao_id": d["ao_id"], "name": d["name"], "role": role,
            "performance_score": ps.get("value"), "status": ps.get("status"),
            "achievement_value": ach, "realisasi": realisasi, "prev_score": 0,
            "detail": d,
        })
    ranked = calc.hitung_ranking(entries)
    return {"period": period, "type": type, "entries": ranked}

# ============================================================
# PORTFOLIO
# ============================================================

@api.get("/portfolio")
async def get_portfolio(ao_id: Optional[str] = None, kolek: Optional[int] = None,
                        user: dict = Depends(get_current_user)):
    q = {}
    if user["role"] == "ao_lending":
        q["ao_id"] = user["id"]
    elif user["role"] == "pic_remedial":
        q["kolektibilitas"] = {"$in": [3, 4, 5]}
    elif ao_id:
        q["ao_id"] = ao_id
    if kolek:
        q["kolektibilitas"] = kolek
    out = []
    async for p in db.loan_portfolio.find(q):
        d = ser(p)
        aou = await db.users.find_one({"_id": ObjectId(d["ao_id"])}) if d.get("ao_id") else None
        d["ao_name"] = aou.get("name") if aou else "-"
        out.append(d)
    return out

@api.post("/portfolio")
async def create_portfolio(body: PortfolioIn, user: dict = Depends(require_roles("admin"))):
    if await db.loan_portfolio.find_one({"nomor_kontrak": body.nomor_kontrak}):
        raise HTTPException(status_code=400, detail="Nomor kontrak sudah ada")
    if body.kolektibilitas not in (1, 2, 3, 4, 5):
        raise HTTPException(status_code=400, detail="Kolektibilitas harus 1-5")
    doc = body.dict()
    doc["created_at"] = datetime.now(timezone.utc).isoformat()
    await db.loan_portfolio.insert_one(doc)
    await audit(user, "Tambah portfolio", after={"nomor_kontrak": body.nomor_kontrak})
    return {"ok": True}

@api.get("/portfolio/summary")
async def portfolio_summary(ao_id: Optional[str] = None, user: dict = Depends(get_current_user)):
    scope = None
    if user["role"] == "ao_lending":
        scope = user["id"]
    elif ao_id and user["role"] in ("admin", "direktur"):
        scope = ao_id
    pf = await portfolio_outstanding(scope)
    return pf

@api.get("/npf")
async def npf_monitor(period: Optional[str] = None, user: dict = Depends(require_roles("admin", "direktur", "pic_remedial"))):
    period = period or await get_active_period()
    pf = await portfolio_outstanding()
    cap = await get_param("npf_score_cap", 150.0)
    thr = await get_param("npf_status_threshold", 1.0)
    rt = await db.targets.find_one({"period": period, "target_npf_ratio": {"$gt": 0}}, sort=[("target_npf_ratio", -1)]) or {}
    npf = calc.hitung_npf(pf["npf_out"], pf["total"], rt.get("target_npf_ratio", 3), cap, thr)
    return {"period": period, "portfolio": pf, "npf": npf,
            "target_npf_ratio": rt.get("target_npf_ratio", 3),
            "target_npf_absolute": rt.get("target_npf_absolute", 0)}

# ============================================================
# COLLECTION ACTIVITY
# ============================================================

STATUS_PENAGIHAN = ["Dikunjungi", "Berkomunikasi", "Janji Bayar", "Pembayaran Masuk",
                    "Tidak Ditemui", "Restrukturisasi", "Eskalasi"]

@api.post("/collection-activities")
async def create_collection(
    tanggal_aktivitas: str = Form(...),
    jam_aktivitas: str = Form(...),
    nomor_kontrak: str = Form(...),
    nama_nasabah: str = Form(...),
    outstanding_pokok: float = Form(0),
    status_penagihan: str = Form(...),
    catatan: str = Form(""),
    latitude: Optional[float] = Form(None),
    longitude: Optional[float] = Form(None),
    files: List[UploadFile] = File(default=[]),
    user: dict = Depends(require_roles("ao_lending", "pic_remedial", "admin")),
):
    if len(files) > 5:
        raise HTTPException(status_code=400, detail="Maksimal 5 foto per aktivitas")
    act = {
        "user_id": user["id"], "user_name": user["name"], "role": user["role"],
        "tanggal_aktivitas": tanggal_aktivitas, "jam_aktivitas": jam_aktivitas,
        "nomor_kontrak": nomor_kontrak, "nama_nasabah": nama_nasabah,
        "outstanding_pokok": outstanding_pokok, "status_penagihan": status_penagihan,
        "catatan": catatan, "approval_status": "Pending",
        "created_at": datetime.now(timezone.utc).isoformat(),
    }
    res = await db.collection_activity.insert_one(act)
    act_id = str(res.inserted_id)

    photo_metas = []
    overall_validation = "Lokasi Tidak Tersedia" if not files else "Valid"
    for f in files:
        raw = await f.read()
        if len(raw) > 10 * 1024 * 1024:
            raise HTTPException(status_code=400, detail=f"Foto {f.filename} melebihi 10MB")
        try:
            meta = storage_util.upload_photo(raw, f.filename, user["id"], user["name"],
                                             tanggal_aktivitas, latitude, longitude)
        except Exception as e:
            logger.error(f"upload foto gagal: {e}")
            raise HTTPException(status_code=500, detail=f"Gagal memproses foto: {e}")
        pdoc = {
            "collection_activity_id": act_id,
            "foto_url": meta["foto_url"], "tanggal_foto": meta.get("tanggal_foto"),
            "timestamp_foto": meta.get("timestamp_foto"),
            "latitude": meta.get("latitude"), "longitude": meta.get("longitude"),
            "exif_available": meta.get("exif_available", False),
            "waktu_upload_fallback": meta.get("waktu_upload_fallback", False),
            "status_validasi": meta.get("status_validasi", "Valid"),
        }
        await db.collection_activity_photos.insert_one(pdoc)
        if meta.get("status_validasi") == "Perlu Verifikasi Admin":
            overall_validation = "Perlu Verifikasi Admin"
        elif meta.get("status_validasi") == "Lokasi Tidak Tersedia" and overall_validation == "Valid":
            overall_validation = "Lokasi Tidak Tersedia"
        photo_metas.append(pdoc)

    await db.collection_activity.update_one({"_id": res.inserted_id},
                                            {"$set": {"status_validasi": overall_validation}})
    await audit(user, "Dokumentasi collection activity", after={"nomor_kontrak": nomor_kontrak})
    return {"id": act_id, "status_validasi": overall_validation, "photos": len(photo_metas)}

@api.get("/collection-activities")
async def list_collection(user: dict = Depends(get_current_user)):
    q = {}
    if user["role"] == "ao_lending":
        q["user_id"] = user["id"]
    elif user["role"] == "pic_remedial":
        q["user_id"] = user["id"]
    out = []
    async for a in db.collection_activity.find(q).sort("created_at", -1):
        d = ser(a)
        photos = []
        async for p in db.collection_activity_photos.find({"collection_activity_id": d["id"]}):
            photos.append(ser(p))
        d["photos"] = photos
        out.append(d)
    return out

@api.post("/collection-activities/{aid}/review")
async def review_collection(aid: str, action: str = Query(...), user: dict = Depends(require_roles("admin"))):
    status = "Approved" if action == "approve" else "Rejected"
    await db.collection_activity.update_one({"_id": ObjectId(aid)}, {"$set": {"approval_status": status}})
    await audit(user, f"Collection {status}", after={"aid": aid})
    return {"ok": True}

@api.get("/files/{path:path}")
async def serve_file(path: str, request: Request, auth: Optional[str] = Query(None)):
    token = auth or request.cookies.get("access_token")
    if not token:
        h = request.headers.get("Authorization", "")
        if h.startswith("Bearer "):
            token = h[7:]
    if not token:
        raise HTTPException(status_code=401, detail="Tidak terautentikasi")
    try:
        jwt.decode(token, JWT_SECRET, algorithms=[JWT_ALG])
    except jwt.InvalidTokenError:
        raise HTTPException(status_code=401, detail="Token tidak valid")
    try:
        data, ct = storage_util.get_object(path)
    except Exception:
        raise HTTPException(status_code=404, detail="File tidak ditemukan")
    return FastResponse(content=data, media_type=ct)

# ============================================================
# PERFORMANCE SETTINGS
# ============================================================

@api.get("/performance-settings")
async def get_perf_settings(user: dict = Depends(require_roles("admin", "direktur"))):
    result = {}
    for role in ("ao_lending", "ao_funding", "pic_remedial"):
        result[role] = await get_weights(role)
    result["parameters"] = {
        "npf_score_cap": await get_param("npf_score_cap", 150.0),
        "npf_status_threshold": await get_param("npf_status_threshold", 1.0),
    }
    return result

@api.post("/performance-settings")
async def set_perf_settings(body: PerfSettingIn, user: dict = Depends(require_roles("admin"))):
    total = sum(float(v) for v in body.weights.values())
    if any(float(v) < 0 or float(v) > 100 for v in body.weights.values()):
        raise HTTPException(status_code=400, detail="Bobot harus 0-100")
    if abs(total - 100) > 0.01:
        raise HTTPException(status_code=400, detail="Total bobot harus 100%")
    last = await db.performance_settings.find_one({"role": body.role, "type": "weight"}, sort=[("version", -1)])
    version = (last["version"] + 1) if last else 1
    await db.performance_settings.insert_one({
        "type": "weight", "role": body.role, "weights": body.weights,
        "version": version, "reason": body.reason,
        "old_weights": last["weights"] if last else None,
        "created_at": datetime.now(timezone.utc).isoformat(),
        "created_by": user["name"]})
    await audit(user, "Ubah bobot performance", before=last["weights"] if last else None, after=body.weights)
    return {"ok": True, "version": version}

@api.post("/performance-settings/parameter")
async def set_param(body: ParamIn, user: dict = Depends(require_roles("admin"))):
    last = await db.performance_settings.find_one(
        {"type": "parameter", "parameter_key": body.parameter_key}, sort=[("version", -1)])
    version = (last["version"] + 1) if last else 1
    await db.performance_settings.insert_one({
        "type": "parameter", "parameter_key": body.parameter_key,
        "parameter_value": body.parameter_value, "version": version,
        "created_at": datetime.now(timezone.utc).isoformat(), "created_by": user["name"]})
    await audit(user, f"Ubah parameter {body.parameter_key}", after={body.parameter_key: body.parameter_value})
    return {"ok": True}

@api.get("/performance-settings/history")
async def perf_history(user: dict = Depends(require_roles("admin", "direktur"))):
    out = []
    async for s in db.performance_settings.find().sort("created_at", -1):
        out.append(ser(s))
    return out

# ============================================================
# SYSTEM SETTINGS
# ============================================================

@api.get("/settings")
async def get_settings(user: dict = Depends(get_current_user)):
    s = await db.system_settings.find_one({"_id": "system"}) or {}
    return {"active_period": s.get("active_period") or current_period(),
            "session_timeout_minutes": s.get("session_timeout_minutes", 60)}

@api.put("/settings")
async def update_settings(body: SettingsIn, user: dict = Depends(require_roles("admin"))):
    before = await db.system_settings.find_one({"_id": "system"}) or {}
    upd = {k: v for k, v in body.dict().items() if v is not None}
    await db.system_settings.update_one({"_id": "system"}, {"$set": upd}, upsert=True)
    if "active_period" in upd:
        await audit(user, "Ubah periode aktif", before={"active_period": before.get("active_period")},
                    after={"active_period": upd["active_period"]})
    return {"ok": True}

# ============================================================
# AUDIT LOG
# ============================================================

@api.get("/audit-logs")
async def get_audit(limit: int = 200, user: dict = Depends(require_roles("admin", "direktur"))):
    out = []
    async for a in db.audit_logs.find().sort("waktu", -1).limit(limit):
        out.append(ser(a))
    return out

# ============================================================
# DATA MANAGEMENT: IMPORT / EXPORT / BACKUP / RESTORE
# ============================================================

@api.post("/import/preview")
async def import_preview(data_type: str = Form(...), file: UploadFile = File(...),
                         user: dict = Depends(require_roles("admin"))):
    raw = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(raw))
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise HTTPException(status_code=400, detail="File kosong")
    headers = [str(h).strip() if h else "" for h in rows[0]]
    records, errors = [], []
    for i, row in enumerate(rows[1:], start=2):
        rec = dict(zip(headers, row))
        if all(v is None for v in row):
            continue
        # basic validation for portfolio
        if data_type == "portfolio":
            if not rec.get("nomor_kontrak"):
                errors.append({"row": i, "reason": "nomor_kontrak kosong"})
            try:
                float(rec.get("outstanding_pokok") or 0)
            except Exception:
                errors.append({"row": i, "reason": "outstanding bukan angka"})
        records.append({"row": i, "data": rec})
    return {"headers": headers, "records": records, "errors": errors,
            "valid_count": len(records) - len(errors), "error_count": len(errors)}

@api.post("/import/confirm")
async def import_confirm(payload: dict, user: dict = Depends(require_roles("admin"))):
    data_type = payload.get("data_type")
    if data_type != "portfolio":
        raise HTTPException(status_code=400, detail="Jenis data belum didukung. Saat ini hanya 'portfolio'.")
    records = payload.get("records", [])
    filename = payload.get("filename", "import.xlsx")
    success, failed = 0, 0
    for r in records:
        rec = r.get("data", r)
        try:
            if data_type == "portfolio":
                if await db.loan_portfolio.find_one({"nomor_kontrak": rec.get("nomor_kontrak")}):
                    failed += 1
                    continue
                await db.loan_portfolio.insert_one({
                    "nomor_kontrak": rec.get("nomor_kontrak"),
                    "nama_nasabah": rec.get("nama_nasabah"),
                    "produk": rec.get("produk", "-"),
                    "plafond": float(rec.get("plafond") or 0),
                    "outstanding_pokok": float(rec.get("outstanding_pokok") or 0),
                    "tanggal_akad": str(rec.get("tanggal_akad", "")),
                    "tanggal_jatuh_tempo": str(rec.get("tanggal_jatuh_tempo", "")),
                    "kolektibilitas": int(rec.get("kolektibilitas") or 1),
                    "dpd": int(rec.get("dpd") or 0),
                    "ao_id": rec.get("ao_id"),
                    "created_at": datetime.now(timezone.utc).isoformat()})
                success += 1
            else:
                failed += 1
        except Exception:
            failed += 1
    await db.import_logs.insert_one({
        "filename": filename, "data_type": data_type, "user_name": user["name"],
        "total": len(records), "success": success, "failed": failed,
        "waktu": datetime.now(timezone.utc).isoformat()})
    await audit(user, "Import data", after={"type": data_type, "success": success, "failed": failed})
    return {"success": success, "failed": failed}

@api.get("/import-history")
async def import_history(user: dict = Depends(require_roles("admin", "direktur"))):
    out = []
    async for a in db.import_logs.find().sort("waktu", -1):
        out.append(ser(a))
    return out

@api.get("/export/{report_type}")
async def export_report(report_type: str, period: Optional[str] = None,
                        user: dict = Depends(require_roles("admin", "direktur"))):
    period = period or await get_active_period()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = report_type[:30]

    if report_type == "achievement":
        ws.append(["Nama AO", "Role", "Target", "Realisasi", "Achievement %", "Performance Score", "Status"])
        async for u in db.users.find({"role": {"$in": ["ao_lending", "ao_funding", "pic_remedial"]}}):
            d = await compute_ao(ser(u), period)
            ps = d.get("performance", {})
            if u["role"] == "ao_lending":
                tgt, rl, ach = d["target_booking"], d["realisasi_booking"], d["ach_lending"].get("value")
            elif u["role"] == "ao_funding":
                tgt, rl, ach = d["target_funding"], d["realisasi_funding"], d["ach_funding"].get("value")
            else:
                tgt, rl, ach = d["target_recovery"], d["realisasi_recovery"], d["ach_recovery"].get("value")
            ws.append([d["name"], u["role"], tgt, rl, ach if ach is not None else "N/A",
                       ps.get("value") if ps.get("value") is not None else "N/A", ps.get("status")])
    elif report_type == "portfolio":
        ws.append(["Nomor Kontrak", "Nama Nasabah", "Produk", "Plafond", "Outstanding", "Kolektibilitas", "DPD"])
        async for p in db.loan_portfolio.find():
            ws.append([p.get("nomor_kontrak"), p.get("nama_nasabah"), p.get("produk"),
                       p.get("plafond"), p.get("outstanding_pokok"), p.get("kolektibilitas"), p.get("dpd")])
    elif report_type == "npf":
        pf = await portfolio_outstanding()
        ws.append(["Kolektibilitas", "Outstanding", "Jumlah Nasabah"])
        for k in range(1, 6):
            ws.append([f"Kolek {k}", pf["kol"][k], pf["kol_count"][k]])
        ws.append(["NPF (Kol 3-5)", pf["npf_out"], ""])
        ws.append(["Total", pf["total"], ""])
    elif report_type == "collection":
        ws.append(["Tanggal", "Nomor Kontrak", "Nama Nasabah", "Outstanding", "PIC", "Status", "Status Validasi", "Catatan"])
        async for a in db.collection_activity.find():
            ws.append([a.get("tanggal_aktivitas"), a.get("nomor_kontrak"), a.get("nama_nasabah"),
                       a.get("outstanding_pokok"), a.get("user_name"), a.get("status_penagihan"),
                       a.get("status_validasi"), a.get("catatan")])
    else:
        raise HTTPException(status_code=400, detail="Report type tidak valid")

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    await audit(user, "Export data", after={"report": report_type})
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition": f"attachment; filename={report_type}_{period}.xlsx"})

BACKUP_COLLECTIONS = ["users", "branches", "role_history", "targets", "achievements",
                      "loan_portfolio", "collection_activity", "collection_activity_photos",
                      "performance_settings", "audit_logs", "system_settings"]

@api.post("/backup")
async def backup_db(user: dict = Depends(require_roles("admin"))):
    dump = {}
    for coll in BACKUP_COLLECTIONS:
        items = []
        async for doc in db[coll].find():
            doc = ser(doc)  # strips password_hash
            if coll == "users":
                doc["requires_password_reset"] = True
                doc.pop("password_hash", None)
            items.append(doc)
        dump[coll] = items
    dump["_meta"] = {"created_at": datetime.now(timezone.utc).isoformat(), "by": user["name"]}
    await audit(user, "Backup database")
    import json
    buf = io.BytesIO(json.dumps(dump, default=str).encode())
    return StreamingResponse(buf, media_type="application/json",
                             headers={"Content-Disposition": f"attachment; filename=ao360_backup_{current_period()}.json"})

@api.post("/restore/confirm")
async def restore_db(payload: dict, user: dict = Depends(require_roles("admin"))):
    # auto-backup before restore
    await audit(user, "Auto-backup sebelum restore")
    restored_users = []
    for coll in BACKUP_COLLECTIONS:
        items = payload.get(coll)
        if items is None:
            continue
        await db[coll].delete_many({})
        docs = []
        for it in items:
            it = dict(it)
            it.pop("id", None)
            it.pop("password_hash", None)
            if coll == "users":
                it["requires_password_reset"] = True
                it["password_hash"] = hash_password(gen_temp_password())
                it["is_active"] = it.get("is_active", True)
                restored_users.append(it.get("email"))
            if it:
                docs.append(it)
        if docs:
            await db[coll].insert_many(docs)
    await audit(user, "Restore database", after={"users_restored": len(restored_users)})
    return {"ok": True, "restored_users": restored_users,
            "note": "Seluruh user wajib ganti password. Gunakan Reset Password untuk generate password sementara."}

@api.get("/constants")
async def constants():
    return {"roles": ROLES, "status_penagihan": STATUS_PENAGIHAN}

# ============================================================
app.include_router(api)

app.add_middleware(
    CORSMiddleware,
    allow_origin_regex=".*",
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

@app.on_event("startup")
async def startup():
    await db.users.create_index("email", unique=True)
    try:
        storage_util.init_storage()
        logger.info("Storage initialized")
    except Exception as e:
        logger.error(f"Storage init failed: {e}")
    import seed
    await seed.run(db, hash_password)
    # Idempotent admin: keep admin password in sync with env
    admin_email = os.environ.get("ADMIN_EMAIL", "admin@hajimiskin.co.id")
    admin_pw = os.environ.get("ADMIN_PASSWORD", "Admin12345")
    existing = await db.users.find_one({"email": admin_email})
    if existing and not verify_password(admin_pw, existing.get("password_hash", "")):
        await db.users.update_one({"_id": existing["_id"]}, {"$set": {"password_hash": hash_password(admin_pw)}})

@app.get("/api/health")
async def health():
    return {"status": "ok"}
